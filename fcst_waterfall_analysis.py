import re
from datetime import datetime

import numpy as np
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# =====================================================
# 0. 기본 설정
# =====================================================

INPUT_FILE = "/kaggle/working/fcst_waterfall_12weeks_sample.xlsx"
OUTPUT_FILE = "/kaggle/working/fcst_waterfall_analysis.xlsx"

CHANGE_RATE_THRESHOLD = 0.20
CHANGE_QTY_THRESHOLD = 100

REQUIRED_COLUMNS = [
    "fcst_issue_week",
    "plan_week",
    "customer",
    "model",
    "product_code",
    "qty",
]

# True면 W01~W52 전체 표시
# False면 입력 데이터에 존재하는 plan_week 범위만 표시, 예: W22~W34
USE_FULL_YEAR_WEEK_COLUMNS = False


# =====================================================
# 1. 주차 변환 함수
# =====================================================

def parse_week(value):
    """
    2026-W22, 2026-W02, W22, W2, 22W, 202622 같은 값을
    year, week_no, sort_key, label 형태로 변환한다.

    핵심 수정:
    - W22를 W02로 잘못 읽지 않도록 두 자리 주차를 먼저 매칭한다.
    """

    if pd.isna(value):
        return np.nan, np.nan, np.nan, np.nan

    text = str(value).strip().upper()

    if text == "":
        return np.nan, np.nan, np.nan, np.nan

    # 연도 추출
    year_match = re.search(r"(19\d{2}|20\d{2})", text)
    year = int(year_match.group(1)) if year_match else np.nan

    # 중요:
    # 두 자리 주차를 먼저 매칭해야 함.
    # 기존 ([1-9]|[1-4]\d|5[0-3]) 순서는 W22를 W2로 잘못 잡음.
    week_match = re.search(r"W0?([1-4]\d|5[0-3]|[1-9])", text)

    # 22W, 1W 형태
    if week_match is None:
        week_match = re.search(r"0?([1-4]\d|5[0-3]|[1-9])W", text)

    # 202622 형태 또는 2026-W22에서 연도 제거 후 숫자 추출
    if week_match is None:
        text_without_year = re.sub(r"(19\d{2}|20\d{2})", "", text)
        nums = re.findall(r"\d+", text_without_year)

        week_no = np.nan
        for num in nums:
            candidate = int(num)
            if 1 <= candidate <= 53:
                week_no = candidate
                break
    else:
        week_no = int(week_match.group(1))

    if pd.isna(week_no):
        return np.nan, np.nan, np.nan, np.nan

    if pd.isna(year):
        year = 9999

    sort_key = int(year) * 100 + int(week_no)
    label = f"W{int(week_no):02d}"
    std = f"{int(year)}-W{int(week_no):02d}" if int(year) != 9999 else label

    return int(year), int(week_no), sort_key, label

def add_week_columns(df, source_col):
    parsed = df[source_col].apply(parse_week)

    df[f"{source_col}_year"] = parsed.apply(lambda x: x[0])
    df[f"{source_col}_week_no"] = parsed.apply(lambda x: x[1])
    df[f"{source_col}_sort_key"] = parsed.apply(lambda x: x[2])
    df[f"{source_col}_label"] = parsed.apply(lambda x: x[3])
    df[f"{source_col}_std"] = parsed.apply(
        lambda x: f"{int(x[0])}-W{int(x[1]):02d}"
        if pd.notna(x[0]) and pd.notna(x[1]) and int(x[0]) != 9999
        else np.nan
    )

    return df


# =====================================================
# 2. 데이터 읽기 및 검증
# =====================================================

def read_input_data():
    df = pd.read_excel(INPUT_FILE)

    missing_cols = [col for col in REQUIRED_COLUMNS if col not in df.columns]

    if missing_cols:
        raise ValueError(f"필수 컬럼이 없습니다: {', '.join(missing_cols)}")

    df = df[REQUIRED_COLUMNS].copy()

    for col in ["fcst_issue_week", "plan_week", "customer", "model", "product_code"]:
        df[col] = df[col].astype(str).str.strip()

    df["qty"] = pd.to_numeric(df["qty"], errors="coerce").fillna(0)

    df = add_week_columns(df, "fcst_issue_week")
    df = add_week_columns(df, "plan_week")

    failed = df[
        df["fcst_issue_week_std"].isna()
        | df["plan_week_std"].isna()
    ]

    if len(failed) > 0:
        print("[경고] 주차 변환 실패 행이 있습니다. 아래 행은 제외됩니다.")
        print(failed[REQUIRED_COLUMNS].head(20))

    df = df[
        df["fcst_issue_week_std"].notna()
        & df["plan_week_std"].notna()
    ].copy()

    return df


# =====================================================
# 3. 중복 데이터 합산
# =====================================================

def aggregate_data(df):
    group_cols = [
        "fcst_issue_week_std",
        "fcst_issue_week_week_no",
        "fcst_issue_week_sort_key",
        "fcst_issue_week_label",
        "plan_week_std",
        "plan_week_week_no",
        "plan_week_sort_key",
        "plan_week_label",
        "customer",
        "model",
        "product_code",
    ]

    result = (
        df.groupby(group_cols, as_index=False, dropna=False)["qty"]
        .sum()
    )

    return result


# =====================================================
# 4. 주차 컬럼 생성
# =====================================================

def make_week_columns(df):
    if USE_FULL_YEAR_WEEK_COLUMNS:
        return [f"W{i:02d}" for i in range(1, 53)]

    min_week = int(df["plan_week_week_no"].min())
    max_week = int(df["plan_week_week_no"].max())

    return [f"W{i:02d}" for i in range(min_week, max_week + 1)]


# =====================================================
# 5. FCST Waterfall View 생성
# =====================================================

def create_waterfall_view(df, value_col, week_columns):
    pivot = pd.pivot_table(
        df,
        index=[
            "customer",
            "model",
            "product_code",
            "fcst_issue_week_std",
            "fcst_issue_week_sort_key",
        ],
        columns="plan_week_label",
        values=value_col,
        aggfunc="sum",
    ).reset_index()

    pivot = pivot.rename(columns={"fcst_issue_week_std": "fcst_issue_week"})

    for week_col in week_columns:
        if week_col not in pivot.columns:
            pivot[week_col] = np.nan

    result = pivot[
        [
            "customer",
            "model",
            "product_code",
            "fcst_issue_week",
            "fcst_issue_week_sort_key",
        ]
        + week_columns
    ].copy()

    result = result.sort_values(
        ["customer", "model", "product_code", "fcst_issue_week_sort_key"]
    ).reset_index(drop=True)

    result = result.drop(columns=["fcst_issue_week_sort_key"])

    return result


# =====================================================
# 6. 전주 대비 FCST 변동 상세 계산
# =====================================================

def create_change_detail(df):
    work = df.sort_values(
        [
            "customer",
            "model",
            "product_code",
            "plan_week_sort_key",
            "fcst_issue_week_sort_key",
        ]
    ).copy()

    compare_keys = [
        "customer",
        "model",
        "product_code",
        "plan_week_sort_key",
    ]

    work["previous_qty"] = work.groupby(compare_keys)["qty"].shift(1)
    work["previous_issue_week"] = work.groupby(compare_keys)["fcst_issue_week_std"].shift(1)

    work["current_qty"] = work["qty"]

    has_previous = work["previous_qty"].notna()

    # 중요:
    # New는 변동이 아니라 신규 기준값이므로 change_qty는 빈값 처리
    work["change_qty"] = np.where(
        has_previous,
        work["current_qty"] - work["previous_qty"],
        np.nan,
    )

    work["change_rate"] = np.where(
        has_previous & (work["previous_qty"] != 0),
        work["change_qty"] / work["previous_qty"],
        np.nan,
    )

    work["change_direction"] = np.select(
        [
            ~has_previous,
            work["change_qty"] > 0,
            work["change_qty"] < 0,
            work["change_qty"] == 0,
        ],
        [
            "New",
            "Increase",
            "Decrease",
            "No Change",
        ],
        default="No Change",
    )

    work["change_flag"] = np.where(
        has_previous
        & (
            (pd.Series(work["change_rate"]).abs() >= CHANGE_RATE_THRESHOLD)
            | (pd.Series(work["change_qty"]).abs() >= CHANGE_QTY_THRESHOLD)
        ),
        True,
        False,
    )

    result = work[
        [
            "fcst_issue_week_std",
            "plan_week_std",
            "customer",
            "model",
            "product_code",
            "current_qty",
            "previous_issue_week",
            "previous_qty",
            "change_qty",
            "change_rate",
            "change_direction",
            "change_flag",
            "fcst_issue_week_sort_key",
            "fcst_issue_week_label",
            "plan_week_sort_key",
            "plan_week_label",
        ]
    ].copy()

    result = result.rename(
        columns={
            "fcst_issue_week_std": "fcst_issue_week",
            "plan_week_std": "plan_week",
        }
    )

    result = result.sort_values(
        [
            "customer",
            "model",
            "product_code",
            "fcst_issue_week_sort_key",
            "plan_week_sort_key",
        ]
    ).reset_index(drop=True)

    return result


# =====================================================
# 7. Change Waterfall View 생성
# =====================================================

def create_change_waterfall_view(change_detail, week_columns):
    change_for_waterfall = change_detail.rename(
        columns={
            "fcst_issue_week": "fcst_issue_week_std",
            "plan_week": "plan_week_std",
        }
    ).copy()

    needed_cols = [
        "customer",
        "model",
        "product_code",
        "fcst_issue_week_std",
        "fcst_issue_week_sort_key",
        "plan_week_label",
        "change_qty",
    ]

    change_for_waterfall = change_for_waterfall[needed_cols].copy()

    result = create_waterfall_view(
        change_for_waterfall,
        value_col="change_qty",
        week_columns=week_columns,
    )

    return result


# =====================================================
# 8. 변동성 상위 품목
# =====================================================

def create_top_volatility_items(change_detail):
    base = change_detail[
        change_detail["change_direction"].isin(["Increase", "Decrease"])
    ].copy()

    columns = [
        "customer",
        "model",
        "product_code",
        "change_count",
        "increase_count",
        "decrease_count",
        "total_abs_change_qty",
        "avg_abs_change_rate",
        "max_abs_change_rate",
        "volatility_rank",
    ]

    if len(base) == 0:
        return pd.DataFrame(columns=columns)

    base["abs_change_qty"] = base["change_qty"].abs()
    base["abs_change_rate"] = base["change_rate"].abs()

    result = (
        base.groupby(["customer", "model", "product_code"], as_index=False)
        .agg(
            change_count=("change_qty", "count"),
            increase_count=("change_direction", lambda x: (x == "Increase").sum()),
            decrease_count=("change_direction", lambda x: (x == "Decrease").sum()),
            total_abs_change_qty=("abs_change_qty", "sum"),
            avg_abs_change_rate=("abs_change_rate", "mean"),
            max_abs_change_rate=("abs_change_rate", "max"),
        )
    )

    result["avg_abs_change_rate"] = result["avg_abs_change_rate"].fillna(0)
    result["max_abs_change_rate"] = result["max_abs_change_rate"].fillna(0)

    result = result.sort_values(
        ["change_count", "total_abs_change_qty", "max_abs_change_rate"],
        ascending=[False, False, False],
    ).reset_index(drop=True)

    result["volatility_rank"] = range(1, len(result) + 1)

    return result[columns]


# =====================================================
# 9. 최신 FCST vs 직전 FCST 비교
# =====================================================

def create_latest_vs_previous(df):
    """
    가장 최근 fcst_issue_week와 직전 fcst_issue_week만 비교한다.
    rolling FCST 구조에서는 최신 발행분이 커버하는 plan_week만 비교 대상이 된다.
    """

    issue_keys = sorted(df["fcst_issue_week_sort_key"].dropna().unique())

    columns = [
        "customer",
        "model",
        "product_code",
        "plan_week",
        "latest_issue_week",
        "previous_issue_week",
        "latest_qty",
        "previous_qty",
        "change_qty",
        "change_rate",
        "change_direction",
        "change_flag",
    ]

    if len(issue_keys) == 0:
        return pd.DataFrame(columns=columns), None, None

    latest_key = issue_keys[-1]
    previous_key = issue_keys[-2] if len(issue_keys) >= 2 else None

    latest_issue_week = df.loc[
        df["fcst_issue_week_sort_key"] == latest_key,
        "fcst_issue_week_std",
    ].iloc[0]

    previous_issue_week = None

    latest = df[df["fcst_issue_week_sort_key"] == latest_key][
        [
            "customer",
            "model",
            "product_code",
            "plan_week_std",
            "plan_week_sort_key",
            "qty",
        ]
    ].copy()

    latest = latest.rename(columns={"qty": "latest_qty"})

    if previous_key is None:
        latest["previous_qty"] = np.nan
        previous_issue_week = np.nan
    else:
        previous_issue_week = df.loc[
            df["fcst_issue_week_sort_key"] == previous_key,
            "fcst_issue_week_std",
        ].iloc[0]

        previous = df[df["fcst_issue_week_sort_key"] == previous_key][
            [
                "customer",
                "model",
                "product_code",
                "plan_week_std",
                "plan_week_sort_key",
                "qty",
            ]
        ].copy()

        previous = previous.rename(columns={"qty": "previous_qty"})

        latest = latest.merge(
            previous,
            on=[
                "customer",
                "model",
                "product_code",
                "plan_week_std",
                "plan_week_sort_key",
            ],
            how="left",
        )

    latest["latest_issue_week"] = latest_issue_week
    latest["previous_issue_week"] = previous_issue_week

    has_previous = latest["previous_qty"].notna()

    latest["change_qty"] = np.where(
        has_previous,
        latest["latest_qty"] - latest["previous_qty"],
        np.nan,
    )

    latest["change_rate"] = np.where(
        has_previous & (latest["previous_qty"] != 0),
        latest["change_qty"] / latest["previous_qty"],
        np.nan,
    )

    latest["change_direction"] = np.select(
        [
            ~has_previous,
            latest["change_qty"] > 0,
            latest["change_qty"] < 0,
            latest["change_qty"] == 0,
        ],
        [
            "New",
            "Increase",
            "Decrease",
            "No Change",
        ],
        default="No Change",
    )

    latest["change_flag"] = np.where(
        has_previous
        & (
            (pd.Series(latest["change_rate"]).abs() >= CHANGE_RATE_THRESHOLD)
            | (pd.Series(latest["change_qty"]).abs() >= CHANGE_QTY_THRESHOLD)
        ),
        True,
        False,
    )

    latest = latest.rename(columns={"plan_week_std": "plan_week"})

    result = latest[
        [
            "customer",
            "model",
            "product_code",
            "plan_week",
            "latest_issue_week",
            "previous_issue_week",
            "latest_qty",
            "previous_qty",
            "change_qty",
            "change_rate",
            "change_direction",
            "change_flag",
            "plan_week_sort_key",
        ]
    ].copy()

    result = result.sort_values(
        ["customer", "model", "product_code", "plan_week_sort_key"]
    ).reset_index(drop=True)

    result = result.drop(columns=["plan_week_sort_key"])

    return result[columns], latest_issue_week, previous_issue_week
    

# =====================================================
# 10. 고객·모델별 요약
# =====================================================

def create_customer_model_summary(latest_vs_previous):
    """
    latest_vs_previous 기준 고객/모델별 변동 요약
    - New는 변동으로 보지 않음
    - Increase / Decrease만 change_count에 포함
    """

    columns = [
        "customer",
        "model",
        "product_count",
        "total_latest_qty",
        "total_change_qty",
        "total_abs_change_qty",
        "increase_qty",
        "decrease_qty",
        "change_count",
        "avg_change_rate",
    ]

    if len(latest_vs_previous) == 0:
        return pd.DataFrame(columns=columns)

    work = latest_vs_previous.copy()

    work["change_qty_for_sum"] = work["change_qty"].fillna(0)
    work["abs_change_qty"] = work["change_qty_for_sum"].abs()

    work["increase_qty"] = np.where(
        work["change_qty_for_sum"] > 0,
        work["change_qty_for_sum"],
        0,
    )

    work["decrease_qty"] = np.where(
        work["change_qty_for_sum"] < 0,
        work["change_qty_for_sum"],
        0,
    )

    work["is_changed"] = work["change_direction"].isin(["Increase", "Decrease"])

    result = (
        work.groupby(["customer", "model"], as_index=False)
        .agg(
            product_count=("product_code", "nunique"),
            total_latest_qty=("latest_qty", "sum"),
            total_change_qty=("change_qty_for_sum", "sum"),
            total_abs_change_qty=("abs_change_qty", "sum"),
            increase_qty=("increase_qty", "sum"),
            decrease_qty=("decrease_qty", "sum"),
            change_count=("is_changed", "sum"),
            avg_change_rate=("change_rate", "mean"),
        )
    )

    result["avg_change_rate"] = result["avg_change_rate"].fillna(0)

    result = result.sort_values(
        ["total_abs_change_qty", "change_count"],
        ascending=[False, False],
    ).reset_index(drop=True)

    return result[columns]


# =====================================================
# 11. Parameter Info
# =====================================================

def create_parameter_info(raw_df, df, latest_issue_week, previous_issue_week, week_columns):
    issue_count = df["fcst_issue_week_std"].nunique()
    comparison_available = issue_count >= 2

    info = [
        ["분석 실행일", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["입력 파일 경로", INPUT_FILE],
        ["결과 파일 경로", OUTPUT_FILE],
        ["변동률 기준", CHANGE_RATE_THRESHOLD],
        ["변동수량 기준", CHANGE_QTY_THRESHOLD],
        ["원본 데이터 행 수", len(raw_df)],
        ["중복 합산 후 행 수", len(df)],
        ["분석 대상 고객 수", df["customer"].nunique()],
        ["분석 대상 모델 수", df["model"].nunique()],
        ["분석 대상 제품 수", df["product_code"].nunique()],
        ["fcst_issue_week 개수", issue_count],
        ["fcst_issue_week 범위", f"{df['fcst_issue_week_std'].min()} ~ {df['fcst_issue_week_std'].max()}"],
        ["plan_week 범위", f"{df['plan_week_std'].min()} ~ {df['plan_week_std'].max()}"],
        ["Waterfall 표시 주차", f"{week_columns[0]} ~ {week_columns[-1]}"],
        ["가장 최근 fcst_issue_week", latest_issue_week],
        ["직전 fcst_issue_week", previous_issue_week],
        ["전주 대비 비교 가능 여부", comparison_available],
        ["비고", "New는 변동으로 보지 않으며 change_flag=False 처리"],
    ]

    return pd.DataFrame(info, columns=["item", "value"])


# =====================================================
# 12. 검증 로직
# =====================================================

def validate_waterfall_structure(fcst_waterfall_view, week_columns):
    missing_cols = [col for col in week_columns if col not in fcst_waterfall_view.columns]

    if missing_cols:
        raise ValueError(f"Waterfall 주차 컬럼 누락: {', '.join(missing_cols)}")

    print("[검증 완료] Waterfall 주차 컬럼이 정상 생성되었습니다.")

    check_rows = fcst_waterfall_view.head(10)

    for _, row in check_rows.iterrows():
        issue_week = row["fcst_issue_week"]
        issue_week_no = int(issue_week.split("-W")[1])
        issue_label = f"W{issue_week_no:02d}"

        if issue_label not in week_columns:
            continue

        issue_idx = week_columns.index(issue_label)
        before_cols = week_columns[:issue_idx]

        if len(before_cols) > 0:
            has_value_before_issue = row[before_cols].notna().any()

            if has_value_before_issue:
                raise ValueError(
                    f"Waterfall 구조 오류: {issue_week} 행에 발행주차 이전 값이 존재합니다."
                )

    print("[검증 완료] 발행주차 이전 구간은 빈칸으로 유지됩니다.")


def validate_change_calculation(change_detail):
    base = change_detail[change_detail["previous_qty"].notna()].copy()

    if len(base) == 0:
        print("[검증 참고] 비교 가능한 previous_qty가 없어 변동 계산 검증을 건너뜁니다.")
        return

    sample = base.sample(min(5, len(base)), random_state=42)

    for idx, row in sample.iterrows():
        expected_change_qty = row["current_qty"] - row["previous_qty"]

        if row["previous_qty"] != 0:
            expected_change_rate = expected_change_qty / row["previous_qty"]
        else:
            expected_change_rate = np.nan

        if not np.isclose(row["change_qty"], expected_change_qty):
            raise ValueError(f"change_qty 계산 오류 index={idx}")

        if pd.notna(expected_change_rate):
            if not np.isclose(row["change_rate"], expected_change_rate):
                raise ValueError(f"change_rate 계산 오류 index={idx}")

    print("[검증 완료] 변동 계산 샘플 검증이 완료되었습니다.")


# =====================================================
# 13. Excel 서식
# =====================================================

def auto_adjust_column_width(ws, min_width=10, max_width=28):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0

        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def apply_common_format(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    auto_adjust_column_width(ws)


def get_header_map(ws):
    return {
        cell.value: cell.column
        for cell in ws[1]
        if cell.value is not None
    }


def apply_number_formats(ws):
    header_map = get_header_map(ws)

    percent_cols = [
        "change_rate",
        "avg_abs_change_rate",
        "max_abs_change_rate",
        "avg_change_rate",
    ]

    qty_keywords = [
        "qty",
        "count",
        "rank",
        "amount",
    ]

    for header, col_idx in header_map.items():
        col_letter = get_column_letter(col_idx)

        if header in percent_cols:
            for cell in ws[col_letter][1:]:
                cell.number_format = "0.0%"

        if any(keyword in str(header) for keyword in qty_keywords):
            for cell in ws[col_letter][1:]:
                cell.number_format = '#,##0'


def apply_fcst_waterfall_format(ws, week_columns, latest_issue_week):
    latest_row_fill = PatternFill("solid", fgColor="E2F0D9")
    qty_fill = PatternFill("solid", fgColor="DDEBF7")

    header_map = get_header_map(ws)
    issue_col = header_map.get("fcst_issue_week")

    week_col_indexes = [
        header_map[col]
        for col in week_columns
        if col in header_map
    ]

    for row in range(2, ws.max_row + 1):
        issue_value = ws.cell(row=row, column=issue_col).value if issue_col else None

        if issue_value == latest_issue_week:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = latest_row_fill

        for col in week_col_indexes:
            cell = ws.cell(row=row, column=col)

            if cell.value not in [None, ""]:
                cell.fill = qty_fill
                cell.number_format = '#,##0'


def apply_change_waterfall_format(ws, week_columns):
    increase_fill = PatternFill("solid", fgColor="F4CCCC")
    decrease_fill = PatternFill("solid", fgColor="CFE2F3")
    strong_fill = PatternFill("solid", fgColor="FFD966")
    strong_font = Font(bold=True)

    header_map = get_header_map(ws)

    week_col_indexes = [
        header_map[col]
        for col in week_columns
        if col in header_map
    ]

    for row in range(2, ws.max_row + 1):
        for col in week_col_indexes:
            cell = ws.cell(row=row, column=col)

            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'

                if abs(cell.value) >= CHANGE_QTY_THRESHOLD:
                    cell.fill = strong_fill
                    cell.font = strong_font
                elif cell.value > 0:
                    cell.fill = increase_fill
                elif cell.value < 0:
                    cell.fill = decrease_fill


def apply_change_detail_format(ws):
    flag_fill = PatternFill("solid", fgColor="FFF2CC")

    header_map = get_header_map(ws)
    flag_col = header_map.get("change_flag")

    if flag_col is None:
        return

    for row in range(2, ws.max_row + 1):
        flag_value = ws.cell(row=row, column=flag_col).value

        if flag_value is True or str(flag_value).upper() == "TRUE":
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = flag_fill


def apply_top_volatility_format(ws):
    top_fill = PatternFill("solid", fgColor="FCE4D6")

    for row in range(2, min(ws.max_row, 11) + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = top_fill


def apply_excel_format(output_file, week_columns, latest_issue_week):
    wb = load_workbook(output_file)

    for ws in wb.worksheets:
        apply_common_format(ws)
        apply_number_formats(ws)

    if "fcst_waterfall_view" in wb.sheetnames:
        apply_fcst_waterfall_format(
            wb["fcst_waterfall_view"],
            week_columns,
            latest_issue_week,
        )

    if "change_waterfall_view" in wb.sheetnames:
        apply_change_waterfall_format(
            wb["change_waterfall_view"],
            week_columns,
        )

    if "fcst_change_detail" in wb.sheetnames:
        apply_change_detail_format(wb["fcst_change_detail"])

    if "top_volatility_items" in wb.sheetnames:
        apply_top_volatility_format(wb["top_volatility_items"])

    if "latest_vs_previous" in wb.sheetnames:
        apply_change_qty_color_format(wb["latest_vs_previous"])

    if "customer_model_summary" in wb.sheetnames:
        apply_change_qty_color_format(wb["customer_model_summary"])

    if "parameter_info" in wb.sheetnames:
        ws = wb["parameter_info"]
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 65

    wb.save(output_file)

def apply_change_qty_color_format(ws):
    """
    change_qty / total_change_qty 컬럼에 색상 적용
    - 증가: 빨간색 계열
    - 감소: 파란색 계열
    - 기준 초과: 노란색 강조
    """

    increase_fill = PatternFill("solid", fgColor="F4CCCC")  # 연한 빨강
    decrease_fill = PatternFill("solid", fgColor="CFE2F3")  # 연한 파랑
    strong_fill = PatternFill("solid", fgColor="FFD966")    # 노랑 강조
    strong_font = Font(bold=True)

    target_headers = [
        "change_qty",
        "total_change_qty",
    ]

    header_map = get_header_map(ws)

    for header in target_headers:
        if header not in header_map:
            continue

        col_idx = header_map[header]

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)

            if not isinstance(cell.value, (int, float)):
                continue

            cell.number_format = '#,##0'

            if abs(cell.value) >= CHANGE_QTY_THRESHOLD:
                cell.fill = strong_fill
                cell.font = strong_font
            elif cell.value > 0:
                cell.fill = increase_fill
            elif cell.value < 0:
                cell.fill = decrease_fill

# =====================================================
# 14. 메인 실행
# =====================================================

def main():
    raw_df = read_input_data()
    df = aggregate_data(raw_df)
    week_columns = make_week_columns(df)

    fcst_waterfall_view = create_waterfall_view(
        df,
        value_col="qty",
        week_columns=week_columns,
    )

    fcst_change_detail = create_change_detail(df)

    change_waterfall_view = create_change_waterfall_view(
        fcst_change_detail,
        week_columns=week_columns,
    )

    top_volatility_items = create_top_volatility_items(fcst_change_detail)

    latest_vs_previous, latest_issue_week, previous_issue_week = create_latest_vs_previous(df)

    customer_model_summary = create_customer_model_summary(latest_vs_previous)

    parameter_info = create_parameter_info(
        raw_df=raw_df,
        df=df,
        latest_issue_week=latest_issue_week,
        previous_issue_week=previous_issue_week,
        week_columns=week_columns,
    )

    validate_waterfall_structure(fcst_waterfall_view, week_columns)
    validate_change_calculation(fcst_change_detail)

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        fcst_waterfall_view.to_excel(
            writer,
            sheet_name="fcst_waterfall_view",
            index=False,
        )

        change_waterfall_view.to_excel(
            writer,
            sheet_name="change_waterfall_view",
            index=False,
        )

        fcst_change_detail[
            [
                "fcst_issue_week",
                "plan_week",
                "customer",
                "model",
                "product_code",
                "current_qty",
                "previous_issue_week",
                "previous_qty",
                "change_qty",
                "change_rate",
                "change_direction",
                "change_flag",
            ]
        ].to_excel(
            writer,
            sheet_name="fcst_change_detail",
            index=False,
        )

        top_volatility_items.to_excel(
            writer,
            sheet_name="top_volatility_items",
            index=False,
        )

        latest_vs_previous.to_excel(
            writer,
            sheet_name="latest_vs_previous",
            index=False,
        )

        customer_model_summary.to_excel(
            writer,
            sheet_name="customer_model_summary",
            index=False,
        )

        parameter_info.to_excel(
            writer,
            sheet_name="parameter_info",
            index=False,
        )

    apply_excel_format(
        OUTPUT_FILE,
        week_columns=week_columns,
        latest_issue_week=latest_issue_week,
    )

    change_flag_count = int(fcst_change_detail["change_flag"].sum())

    if len(top_volatility_items) > 0:
        top_item = top_volatility_items.iloc[0]
        top_product_text = (
            f"{top_item['customer']} / {top_item['model']} / {top_item['product_code']}"
        )
    else:
        top_product_text = "없음"

    max_increase = fcst_change_detail["change_qty"].max()
    max_decrease = fcst_change_detail["change_qty"].min()

    print()
    print("FCST Waterfall 변동 분석 완료")
    print(f"생성 파일 경로: {OUTPUT_FILE}")
    print(f"원본 데이터 행 수: {len(raw_df):,}")
    print(f"중복 합산 후 행 수: {len(df):,}")
    print(f"분석 대상 고객 수: {df['customer'].nunique():,}")
    print(f"분석 대상 모델 수: {df['model'].nunique():,}")
    print(f"분석 대상 제품 수: {df['product_code'].nunique():,}")
    print(f"fcst_issue_week 범위: {df['fcst_issue_week_std'].min()} ~ {df['fcst_issue_week_std'].max()}")
    print(f"plan_week 범위: {df['plan_week_std'].min()} ~ {df['plan_week_std'].max()}")
    print(f"Waterfall 표시 주차: {week_columns[0]} ~ {week_columns[-1]}")
    print(f"가장 최근 fcst_issue_week: {latest_issue_week}")
    print(f"직전 fcst_issue_week: {previous_issue_week}")
    print(f"변동 기준 초과 건수: {change_flag_count:,}")
    print(f"변동성 1위 제품: {top_product_text}")
    print(f"최대 증가 수량: {max_increase:,.0f}")
    print(f"최대 감소 수량: {max_decrease:,.0f}")


if __name__ == "__main__":
    main()
